
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse

import sqlite3
import datetime

import openpyxl, pyexcel
from .models import Message, Message_User, Notice, Photo, Sender_User
import pandas as pd


from .func.katalk_send import katalk_send

from .forms import UploadFileForm

from django.http import HttpResponse, Http404, HttpResponseRedirect, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt




#-------------------함수 기반 view---------------------
# render 안쓴 것
# def index(request):
#     latest_question_list = Question.objects.order_by('-pub_date')[:5]
#     template = loader.get_template('sms/index_test.html')
#     context = {
#         'latest_question_list': latest_question_list,  # context를 통해 데이터를 전달
#     }
#     return HttpResponse(template.render(context, request))
# render 적용


@csrf_exempt
def index(request):
    if request.method == 'GET':
        conn = sqlite3.connect('./db.sqlite3')
        cur = conn.cursor()
        user_id = request.session.get('user')



        sql = "select * from sms_notice where notice_id_id =?"
        cur.execute(sql, [user_id])
        rows = cur.fetchall()
        rows = pd.DataFrame(rows)


        rows.rename(columns={0:"date", 1:"title",2:"text"},inplace=True)

        rate = []
        notice_pk=[]
        date =[] # 2021. 2. 16. 오전 11:28:54        2021-02-09 01:47:37

        if 0<len(rows):
            for i in rows['date']:
                receive_sum = 0
                date.append(i[:16])

                sql = "select isConfirmbyReceiver from sms_Message where notice_pk_id =?"

                notice_pk.append(i+" "+ user_id)
                cur.execute(sql, [i+" "+ user_id])
                print(i+" "+ user_id)
                row = cur.fetchall()
                print(row)
                for receive in row:

                    receive_sum+=receive[0]
                if receive_sum!=0:
                    rate.append(  round ( (receive_sum/len(row))* 100))
                else:
                    rate.append(0)
                # row = pd.DataFrame(row)




            title = rows['title']
            text = rows['text']



            request.session['user'] = user_id
            ## 공지 보낸 페이지 여기 까지


            ##=========== 유저 목록 데이터 ============

            sql = "select * from sms_Message_User where message_User_id_id =?"
            cur.execute(sql, [user_id])
            rows = cur.fetchall()
            rows = pd.DataFrame(rows)

            rows.rename(columns={0: "user_phoneNumber", 1: "user_name", 2: "user_dong",3:"user_hosu"}, inplace=True)



            user_phoneNumber = rows['user_phoneNumber']
            user_name = rows['user_name']
            user_dong = rows['user_dong']
            user_hosu = rows['user_hosu']

            sql = "select distinct user_dong from sms_Message_User where message_User_id_id =?"
            cur.execute(sql, [user_id])

            rows = cur.fetchall()
            dong_list = []
            group_list = {}


            for i in range(len(rows)):
                dong_list.append(rows[i][0])
                group_list[rows[i][0]]=[]



            sql = "select  user_dong, user_hosu from sms_Message_User where message_User_id_id =?"
            cur.execute(sql, [user_id])

            rows = cur.fetchall()



            for i in range(len(rows)):
                group_list[rows[i][0]].append(rows[i][1])


            history_data = zip(date, title, text,rate,notice_pk)

            user_data = zip(user_phoneNumber, user_name, user_dong,user_hosu,)




            return render(request, 'sms/index.html',{"history_data":history_data, "user_data":user_data,"dong_list":dong_list,"group_list":sorted(group_list.items()),"rate":rate})
        return render(request, 'sms/index.html')
    # conn = sqlite3.connect('./db.sqlite3')
    # cur = conn.cursor()
    # sql = "select distinct user_dong from  main.sms_message_user"
    #
    # cur.execute(sql)
    # rows = cur.fetchall()
    #
    # group_list=[]
    # for i in range(len(rows)):
    #     group_list.append(rows[i][0])
    #
    #
    #
    # return render(request, 'sms/index_test.html',{'kind': group_list})

@csrf_exempt
def insert_notice(request):
    if request.method == 'POST':

        user_id = request.session.get('user')
        request.session['user'] = user_id
        print(user_id)

        return render(request, 'sms/createNotice.html')


@csrf_exempt
def createNotice(request):

    if request.method == 'GET':
        conn = sqlite3.connect('./db.sqlite3')
        cur = conn.cursor()
        user_id = request.session.get('user')
        insert = request.GET.get('insert')



        sql = "select distinct user_dong from sms_Message_User where message_User_id_id =?"
        cur.execute(sql, [user_id])

        rows = cur.fetchall()
        dong_list = []
        group_list = {}
        insert_list = request.GET.getlist('input')
        dong = request.GET.get('group')

        if insert=="등록":



            insert_list.append(dong)
        elif 0!= len(insert_list):
            insert_list.pop(-1)




        for i in range(len(rows)):
            dong_list.append(rows[i][0])
            group_list[rows[i][0]] = []

        sql = "select  user_dong, user_hosu from sms_Message_User where message_User_id_id =?"
        cur.execute(sql, [user_id])

        rows = cur.fetchall()

        for i in range(len(rows)):
            group_list[rows[i][0]].append(rows[i][1])





        return render(request,'sms/createNotice.html',{"dong_list":dong_list,"group_list":sorted(group_list.items()), "insert_list":insert_list})
    if request.method == 'POST':
        # phone_number = request.POST.getlist('phone_number')
        # name = request.POST.getlist('name')

        conn = sqlite3.connect('./db.sqlite3')
        cur = conn.cursor()

        # 공지 DB Insert

        notice_text = request.POST.get('notice_text')
        notice_title = request.POST.get('notice_title')
        user_id = request.session.get('user')
        print(notice_text)

        now = datetime.datetime.now()
        nowDatetime = now.strftime('%Y-%m-%d %H:%M:%S')
        print(nowDatetime)


        # sql = "insert into sms_notice (notice_date,notice_title,notice_text,notice_id_id,notice_pk) values (?,?,?,?,?)"
        #
        # cur.execute(sql, [nowDatetime, str(notice_title), str(notice_text), str(user_id),nowDatetime +" "+str(user_id) ])
        #
        # conn.commit()

        notice = Notice()
        notice.notice_date = nowDatetime
        notice.notice_title = str(notice_title)
        notice.notice_text =str(notice_text)
        notice.notice_id_id = str(user_id)
        notice.notice_pk = nowDatetime +" "+str(user_id)
        notice.save()


        # 알림톡 수신자 data 가져오기
        group = request.POST.get('group')
        insert_list = request.POST.getlist('send')
        my_set = set(insert_list)  # 집합set으로 변환
        insert_list = list(my_set)  # list로 변환


        if "전체"  in insert_list :
            sql = "select * from  main.sms_message_user"
            cur.execute(sql)
        else:
            for i in range(len(insert_list)):

                sql = "select * from  main.sms_message_user where user_dong in (?) "

            cur.execute(sql, insert_list)

        rows = cur.fetchall()

        rows = pd.DataFrame(rows)

        print(rows)
        phone_number_list = rows[0]
        name = rows[1]
        user_dong_hosu_list=rows[4]

        for img in request.FILES.getlist('imgs'):

            #Photo 객체를 하나 생성한다.
            photo = Photo()
            # 외래키로 현재 생성한 Notice의 기본키를 참조한다.
            photo.notice_pk = notice
            # imgs로부터 가져온 이미지 파일 하나를 저장한다.
            photo.image = img
            # 데이터베이스에 저장
            photo.save()

        # sql = "insert into sms_photo (notice_date_id,image) values (?,?)"
        #
        # cur.execute(sql, [nowDatetime, str(img)])
        #
        # conn.commit()
        #
        # cur.close()
        # conn.close()

        # limit_time = request.POST.get('limit_time')
        send = katalk_send(name, phone_number_list,user_dong_hosu_list ,  str(notice_text), str(notice_title), str("12시"), str(user_id),
                           str(nowDatetime))
        send.send()

    return HttpResponseRedirect(
        reverse('sms:index', ))


@csrf_exempt
def noticeDetail(request,notice_pk):
    if request.method == "GET":
        user_id = request.session.get('user')


        conn = sqlite3.connect('./db.sqlite3')
        cur = conn.cursor()
        print(notice_pk)
        sql = "select * from sms_Notice where notice_pk = ?"
        cur.execute(sql,[str(notice_pk)])
        rows = cur.fetchall()

        title = rows[0][1]

        rows = pd.DataFrame(rows)

        rows.rename(columns={0: "date", 1: "title", 2: "text"}, inplace=True)
        day = rows['date']
        rate = []

        receive_sum = 0

        sql = "select isConfirmbyReceiver from sms_Message where notice_pk_id =?"


        cur.execute(sql, [notice_pk])

        row = cur.fetchall()

        ## 수신률 계산
        for receive in row:
            receive_sum += receive[0]
            print(receive[0])
        if receive_sum != 0:
            rate.append(round((receive_sum / len(row)) * 100))
        else:
            rate.append(0)
        # row = pd.DataFrame(row)


        ## 수신률 계산 끝


        ## 메세지  미확인자 확인자 체크
        sql = "select * from sms_Message where notice_pk_id = ?"
        cur.execute(sql,[str(notice_pk)])

        total = {}
        verify={}
        Nverify={}

        rows = cur.fetchall()

        for data in rows:
            sql = "select * from sms_Message_User where user_phoneNumber = ? and  message_User_id_id = ?"
            cur.execute(sql, [str(data[-1]), str(user_id)])

            dong_hosu=data[6].split(" ")

            total[dong_hosu[0]] = dong_hosu[1]  ## 동    호수 저장 딕셔너리


            if data[2]==0:
                Nverify[dong_hosu[0]] =dong_hosu[1]

            else:
                verify[dong_hosu[0]] = dong_hosu[1]


        text =rows[0][0]

        ## 메세지  미확인자 확인자 체크 끝








        return render(request, 'sms/noticeDetail.html',{"title":title,"text":text,"rate":int(rate[0]),"day":str(day[0]) ,
                                                        "total":sorted(total.items()),"total_len":len(total),
                                                        "verify":sorted(verify.items()),
                                                        "verify_len":len(verify),
                                                        "Nverify_len":len(Nverify),"Nverify":sorted(Nverify.items())})
@csrf_exempt
def register(request):

    return render(request, 'sms/register.html')
@csrf_exempt
def register_submit(request):
    try:

        if request.method == 'POST':

            conn = sqlite3.connect('./db.sqlite3')
            cur = conn.cursor()
            sql = "select * from sms_Sender_User where id =?"
            cur.execute(sql,[str(request.POST.get('id'))] )
            rows = cur.fetchall()

            if len(rows)>0:

                return HttpResponse("ID가 중복 되거나 기타 오류 입니다.")



            sender_User = Sender_User()
            sender_User.id=request.POST.get('id')
            sender_User.ps =request.POST.get('ps')
            sender_User.aptNm =request.POST.get('aptNm')
            sender_User.complexNm =request.POST.get('complexNm')
            sender_User.villageNm =request.POST.get('villageNm')
            sender_User.aptAddress =request.POST.get('aptAddress')
            sender_User.save()

            return redirect('/sms/login/')
    except:
        return HttpResponse("ID가 중복 되거나 기타 오류 입니다.")

@csrf_exempt
def upload(request):

    if request.method == "POST":
        user_id = request.session.get('user')


        form = UploadFileForm(request.POST, request.FILES)


        if form.is_valid():

            conn = sqlite3.connect('./db.sqlite3')
            cur = conn.cursor()

            excel_file = request.FILES["file"]
            wb = openpyxl.load_workbook(excel_file)
            # getting a particular sheet by name out of many sheets


            excel_data = list()
            # iterating over the rows and
            # getting value from each cell in row
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            for row in ws.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))

                   # print(str(cell.value), "row_data")
                row_data.append(user_id)
                #print(row_data)
                excel_data.append(row_data)

            del excel_data[0]
            for data in excel_data:
                data.append(data[2]+" "+data[3])
                print(data)

                sql2 ="INSERT INTO sms_Message_User " \
                      "(user_phoneNumber,user_name,user_dong,user_hosu,message_User_id_id ,user_pk)" \
                      " values (?,?,?,?,?,?) ON CONFLICT(user_pk) DO UPDATE SET user_phoneNumber = '"+(data[0])+"',  user_name = '"+(data[1])+"'  "

                cur.execute(sql2, data)

                conn.commit()


                # message_User = Message_User()
                # message_User.user_phoneNumber =
                # message_User.user_name =
                # message_User.user_dong =
                # message_User.user_hosu =
                # message_User.message_User_id =
                # message_User.save()


            # excel_data = list()
            # # iterating over the rows and
            # # getting value from each cell in row
            # for row in worksheet.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data.append(row_data)
            #
            #
            #
            # request.FILES['file'].save_to_database(
            #     model=Message_User,
            #     mapdict=['user_phoneNumber', 'user_name', 'user_dong','user_hosu','message_User_id'])
            request.session['user'] = user_id
            print(user_id)
            return render(request, 'sms/index.html')
        else:

            return HttpResponseBadRequest(form.is_valid())
    else:

        return render(request, 'sms/index_test.html')



@csrf_exempt
def login(request):
    if request.method == 'GET':
        return render(request, 'sms/login.html')

    if request.method == 'POST':
        print("POST")
        # phone_number = request.POST.getlist('phone_number')
        # name = request.POST.getlist('name')

        conn = sqlite3.connect('./db.sqlite3')
        cur = conn.cursor()

        # 공지 DB Insert

        # id = request.POST.get('id')
        # ps = request.POST.get('ps')

        sql = "select * from sms_Sender_User where id =? and ps =?"
        cur.execute(sql, [str(request.POST.get('id')), str(request.POST.get('ps'))])
        rows = cur.fetchall()
        if len(rows) == 1:
            request.session['user'] = request.POST.get('id')
            return redirect('/sms/index/')
        else:
            return HttpResponse("ID가 중복 되거나 없는 ID 입니다.")



    #return render(request, 'sms/login_test.html') ## form

@csrf_exempt
def message_User_insert(request):

    if request.method == "POST":

        form = UploadFileForm(request.POST, request.FILES)


        if form.is_valid():

            request.FILES['file'].save_to_database(
                model=Message_User,
                mapdict=['user_phoneNumber', 'user_name', 'user_group'])
            return HttpResponse("OK")
        else:

            return HttpResponseBadRequest(form.is_valid())
    else:
        return render(request, 'sms/index_test.html')



@csrf_exempt
def m_noticeDetail(request,notice_url):
    if request.method == 'GET':




        conn = sqlite3.connect('./db.sqlite3')
        cur = conn.cursor()

        sql = "select user_pk_id from sms_message WHERE notice_url  = ?"

        cur.execute(sql,[notice_url])
        rows = cur.fetchall()

        now = datetime.datetime.now()
        nowDatetime = now.strftime('%Y-%m-%d %H:%M:%S')

        cur.execute("UPDATE sms_message SET isConfirmbyReceiver = True , notice_Confirm_date = ? WHERE user_pk_id = ? ",[nowDatetime,rows[0][0]])


        message_User =  get_object_or_404(Message_User,pk=rows[0][0])

        message = get_object_or_404(Message,pk=notice_url)






        sql = "select * from sms_notice where notice_date =?"
        cur.execute(sql, [notice_url[:19]])
        rows = cur.fetchall()

        title= rows[0][1]
        sender_id =rows[0][-1]

        # sql = "select  image from main.sms_Photo WHERE notice_pk_id  = ?"
        # # print(notice_url)  ## 2021-01-27 16:17:5001082745538      날짜 pk 랑 휴대전화 나눠야함
        #
        # cur.execute(sql, [rows[0][3]])
        # rows = cur.fetchall()
        #
        # photo=rows[0][0]
        # print(photo)
        notice = get_object_or_404(Notice,pk=rows[0][3])



        conn.commit()
        cur.close()
        conn.close()

        return render(request, 'sms/m_noticeDetail.html',
                      {'message': message, 'message_User': message_User, "title":title,"sender_id":sender_id ,"date":notice_url[:19],'notice':notice})
    else:
        return render(request, 'sms/m_noticeDetail.html')
